from openpyxl import *
import pandas as pd
import xlrd


def write_excel(data, filename, keys=''):
    data_df = pd.DataFrame(data)
    if keys != '':
        data_df.columns = list(keys)
    writer = pd.ExcelWriter('{}.xlsx'.format(filename))
    data_df.to_excel(writer, index=None)
    writer.save()


def write_csv(name, data_list):
    fieldnames = data_list[0]
    print(fieldnames)
    with open(name + '.csv', mode='w', newline='', encoding='utf-8-sig') as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        for data in data_list:
            writer.writerow(data)


def parse_list(data):
    """
    parse a list
    :param data: list
    :return:
    """
    data_list = []
    for i in range(1, len(data)):
        data_dict = {}
        for index, key in enumerate(data[0]):
            data_dict[key] = data[i][index]
        data_list.append(data_dict)
    return data_list


def read_excel(filename, sheet=None):
    """
    read excel and parse every line

    :param name: file of absolute path
    :return: some dict which is about everyline of list
    """
    workbook = xlrd.open_workbook(filename)
    total_data = []
    if sheet:
        sh = workbook.sheet_by_name(sheet)
        for row in range(0, sh.nrows):
            row_list = sh.row_values(row)
            total_data.append(row_list)
    return parse_list(total_data)


class excel():
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    # 获取表格的总行数和总列数
    def getRowsClosNum(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def getCellValue(self, row, column):
        cellvalue = self.ws.cell(row=row, column=column).value
        return cellvalue

    # 获取某列的所有值
    def getColValues(self, column):
        rows = self.ws.max_row
        columndata = []
        for i in range(1, rows + 1):
            cellvalue = self.ws.cell(row=i, column=column).value
            columndata.append(cellvalue)
        return columndata

    # 获取某行所有值
    def getRowValues(self, row):
        columns = self.ws.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    # 设置某个单元格的值
    def setCellValue(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)

    # 获取某个单元格颜色，前两位为透明度，ff为完全不透明，00为完全透明
    def getColor(self, row, col):
        """
        :param row:
        :param col:
        :return: 单元格值，字体颜色，填充颜色（背景色）
        """
        cell = self.ws.cell(row=row, column=col)
        try:
            return cell.value, cell.font.color.rgb, cell.fill.start_color.rgb
        except:
            return cell.value, ''
